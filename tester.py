from openpyxl import load_workbook
import pandas as pd
wb2 = load_workbook(r'C:\Users\nailt\Desktop\Machine-Learning-for-Solar-Energy-Prediction-master\staj-final\excel-files\kaski\demokrasi.xlsx')
df = pd.DataFrame(wb2.values)
import openpyxl
print(openpyxl.__version__)
sheet_names = wb2.get_sheet_names()
name = sheet_names[0]
sheet_ranges = wb2[name]
df = pd.DataFrame(sheet_ranges.values)
df.columns = df.iloc[0]
df = df.iloc[1:]
df
def xlsx_2_df(path, first_row_header = True):
    wb = openpyxl.load_workbook(path)
    sheet_names = wb.get_sheet_names()
    df = pd.DataFrame()
    for i in range (len(sheet_names)):
        carrier = sheet_names[i]
        sheet_ranges = wb[carrier]
        df_carrier = pd.DataFrame(sheet_ranges.values)
        df_carrier.columns = df_carrier.iloc[0]
        df_carrier = df_carrier.iloc[1:]
        df.append(df_carrier)
    return df
path = r'C:\Users\nailt\Desktop\Machine-Learning-for-Solar-Energy-Prediction-master\staj-final\excel-files\kaski\demokrasi.xlsx'
xlsx_2_df(path)
